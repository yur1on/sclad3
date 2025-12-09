
from django.db.models import Value, IntegerField, F
from django.db.models.functions import Lower, Replace
from django.db.models import Case, When, Q
from django.contrib.auth import logout
from pathlib import Path
import json
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse
from itertools import groupby
from django.db.models import Q, OuterRef, Subquery
from django.core.paginator import Paginator
from .telegram_utils import delete_telegram_message_for_part
import os
from django.conf import settings
from django.db.models import F
from django.utils import timezone
from datetime import timedelta
from django.db.models import Sum, DateField
from django.db.models.functions import TruncDate
from .forms import PartForm, PartImageFormSet
from tariff.utils import check_parts_limit, check_image_limit
from .utils import compress_image
from warehouse.utils import add_watermark_to_image
from .telegram_utils import send_new_part_notification
from django.utils.timezone import now
from django.db.models import Sum
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db import transaction
from django.contrib import messages
from .models import Part, PartImage
from decimal import Decimal, InvalidOperation
from django.db.models import Count, Min, Max



def home(request):
    # последние запчасти с количеством > 0
    results = (
        Part.objects.filter(quantity__gt=0)
        .select_related("user", "user__profile")
        .prefetch_related("images")
        .order_by("-created_at")
    )

    # По желанию — ограничение/приоритизация как в search():
    now = timezone.now()
    active_paid = Q(user__profile__tariff__in=['lite','standard','standard2','standard3','premium']) & \
                  Q(user__profile__subscription_end__isnull=False, user__profile__subscription_end__gte=now)
    subquery = Part.objects.filter(user=OuterRef('user_id'), quantity__gt=0).order_by('-created_at').values('pk')[:30]
    results = results.filter(active_paid | Q(pk__in=subquery))

    paginator = Paginator(results, 30)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "page_obj": page_obj,
        # "advertisements": Advertisement.objects.all()[:4],  # если есть
        "advertisements": [],
        "query": "",  # для безопасной подстановки в инпут
    }
    return render(request, "warehouse/home.html", context)


from decimal import Decimal, InvalidOperation

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import (
    Q, F, Value, IntegerField, OuterRef, Subquery,
    Count, Min, Max, Sum, Case, When
)
from django.db.models.functions import Lower, Replace
from django.shortcuts import render
from django.utils import timezone

from .models import Part
from .search_utils import split_search_q, note_regex_from_text, normalize_compact


def _annotate_search_fields(qs):
    """
    Общие аннотации для поиска.
    normalized_* убирают пробелы и дефисы (чтобы SM-A536B находилось по A536B).
    """
    return qs.annotate(
        lower_brand=Lower(F("brand")),
        lower_model=Lower(F("model")),
        lower_chip_label=Lower(F("chip_label")),

        # model: убираем пробелы и дефисы
        normalized_model=Lower(
            Replace(
                Replace(F("model"), Value(" "), Value("")),
                Value("-"), Value("")
            )
        ),

        # part_number: ТОЛЬКО ПОЛНОЕ совпадение (тоже нормализуем пробелы/дефисы)
        normalized_part_number=Lower(
            Replace(
                Replace(F("part_number"), Value(" "), Value("")),
                Value("-"), Value("")
            )
        ),
    )


def search(request):
    q_raw = (request.GET.get("q") or "").strip()

    # совместимость со старыми ссылками
    brand_in = (request.GET.get("brand") or "").strip()
    model_in = (request.GET.get("model") or "").strip()

    brand, model, single_term, term, q = split_search_q(q_raw, brand_in, model_in)

    # -------- Фильтры UI --------
    part_types = [p.strip() for p in request.GET.getlist("part_type") if p.strip()]
    region = (request.GET.get("region") or "").strip()
    city = (request.GET.get("city") or "").strip()
    condition = (request.GET.get("condition") or "").strip()
    color = (request.GET.get("color") or "").strip()
    sort = (request.GET.get("sort") or "relevance").strip()

    price_min_raw = (request.GET.get("price_min") or "").strip()
    price_max_raw = (request.GET.get("price_max") or "").strip()
    has_images = (request.GET.get("has_images") == "1")

    # -------- База queryset --------
    results = (
        Part.objects
        .filter(quantity__gt=0)
        .select_related("user", "user__profile")
        .prefetch_related("images")
    )
    results = _annotate_search_fields(results)

    brand_l = (brand or "").strip().lower()
    model_l = (model or "").strip().lower()

    term_l = (term or "").strip().lower()

    model_norm = normalize_compact(model)
    term_norm = normalize_compact(term)

    # note: только модельные токены (с цифрами) и как отдельные слова
    note_pat_term = note_regex_from_text(term) if term else None
    note_pat_model = note_regex_from_text(model) if model else None

    # -------- Поиск + приоритет --------
    if single_term and term:
        base_q = (
            # model / chip
            Q(normalized_model__icontains=term_norm) |
            Q(lower_model__icontains=term_l) |
            Q(lower_chip_label__icontains=term_l) |
            # part_number — ТОЛЬКО полное
            Q(normalized_part_number=term_norm) |
            # brand
            Q(lower_brand__icontains=term_l)
        )
        if note_pat_term:
            base_q = base_q | Q(note__iregex=note_pat_term)

        results = results.filter(base_q)

        whens = [
            When(normalized_part_number=term_norm, then=Value(120)),     # part_number full
            When(normalized_model=term_norm, then=Value(95)),            # model exact (compact)
            When(lower_model=term_l, then=Value(90)),                    # model exact text
            When(normalized_model__icontains=term_norm, then=Value(75)), # model contains
            When(lower_model__icontains=term_l, then=Value(70)),
            When(lower_chip_label__icontains=term_l, then=Value(60)),
            When(lower_brand__icontains=term_l, then=Value(40)),
        ]
        if note_pat_term:
            whens.append(When(note__iregex=note_pat_term, then=Value(25)))  # note ниже модели/бренда

        results = results.annotate(
            match_priority=Case(*whens, default=Value(0), output_field=IntegerField())
        )

    else:
        # бренд (если есть)
        if brand:
            results = results.filter(lower_brand__icontains=brand_l)

        if model:
            model_q = (
                Q(normalized_model__icontains=model_norm) |
                Q(lower_model__icontains=model_l) |
                Q(lower_chip_label__icontains=model_l) |
                Q(normalized_part_number=model_norm)   # part_number full
            )
            if note_pat_model:
                model_q = model_q | Q(note__iregex=note_pat_model)

            results = results.filter(model_q)

        whens = []
        if model:
            whens.extend([
                When(normalized_part_number=model_norm, then=Value(115)),  # part_number full
                When(lower_brand=brand_l, normalized_model=model_norm, then=Value(100)),
                When(lower_brand=brand_l, lower_model=model_l, then=Value(95)),
                When(normalized_model=model_norm, then=Value(85)),
                When(lower_model=model_l, then=Value(80)),
                When(normalized_model__icontains=model_norm, then=Value(65)),
                When(lower_model__icontains=model_l, then=Value(60)),
                When(lower_chip_label__icontains=model_l, then=Value(55)),
            ])
            if note_pat_model:
                whens.append(When(note__iregex=note_pat_model, then=Value(25)))

        if brand:
            whens.append(When(lower_brand=brand_l, then=Value(30)))

        results = results.annotate(
            match_priority=Case(*whens, default=Value(0), output_field=IntegerField())
        )

    # -------- Ограничения подписки (как у тебя было) --------
    now = timezone.now()
    active_paid = (
        Q(user__profile__tariff__in=["lite", "standard", "standard2", "standard3", "premium"]) &
        Q(user__profile__subscription_end__isnull=False, user__profile__subscription_end__gte=now)
    )
    last_30_pk = (
        Part.objects
        .filter(user=OuterRef("user_id"), quantity__gt=0)
        .order_by("-created_at")
        .values("pk")[:30]
    )
    results = results.filter(active_paid | Q(pk__in=Subquery(last_30_pk)))

    # -------- Фильтры --------
    if part_types:
        results = results.filter(part_type__in=part_types)
    if region:
        results = results.filter(user__profile__region__icontains=region)
    if city:
        results = results.filter(user__profile__city__icontains=city)
    if condition:
        results = results.filter(condition__icontains=condition)
    if color:
        results = results.filter(color__icontains=color)
    if has_images:
        results = results.filter(images__isnull=False).distinct()

    try:
        if price_min_raw:
            results = results.filter(price__gte=Decimal(price_min_raw.replace(",", ".")))
    except (InvalidOperation, ValueError):
        pass

    try:
        if price_max_raw:
            results = results.filter(price__lte=Decimal(price_max_raw.replace(",", ".")))
    except (InvalidOperation, ValueError):
        pass

    # -------- Фасеты --------
    facets_base = results

    facet_part_types = list(
        facets_base.exclude(part_type__isnull=True).exclude(part_type="")
        .values("part_type").annotate(count=Count("id")).order_by("part_type")
    )
    facet_conditions = list(
        facets_base.exclude(condition__isnull=True).exclude(condition="")
        .values("condition").annotate(count=Count("id")).order_by("condition")
    )
    facet_colors = list(
        facets_base.exclude(color__isnull=True).exclude(color="")
        .values("color").annotate(count=Count("id")).order_by("color")
    )
    facet_cities = list(
        facets_base.exclude(user__profile__city__isnull=True).exclude(user__profile__city="")
        .values(city=F("user__profile__city")).annotate(count=Count("id")).order_by("city")
    )
    facet_regions = list(
        facets_base.exclude(user__profile__region__isnull=True).exclude(user__profile__region="")
        .values(region=F("user__profile__region")).annotate(count=Count("id")).order_by("region")
    )
    price_bounds = facets_base.aggregate(min_price=Min("price"), max_price=Max("price"))

    # -------- Сортировка --------
    if sort == "new":
        results = results.order_by("-created_at")
    elif sort == "price_asc":
        results = results.order_by("price", "-match_priority", "-created_at")
    elif sort == "price_desc":
        results = results.order_by("-price", "-match_priority", "-created_at")
    elif sort == "qty_desc":
        results = results.order_by("-quantity", "-match_priority", "-created_at")
    else:
        results = results.order_by("-match_priority", "-created_at")

    # -------- Пагинация --------
    total_results = results.count()
    paginator = Paginator(results, 30)
    page_obj = paginator.get_page(request.GET.get("page"))

    qs = request.GET.copy()
    qs.pop("page", None)
    base_qs = qs.urlencode()

    has_filters = any([
        part_types, region, city, condition, color, price_min_raw, price_max_raw, has_images,
        (sort and sort != "relevance")
    ])

    return render(request, "warehouse/search.html", {
        "page_obj": page_obj,
        "total_results": total_results,
        "base_qs": base_qs,
        "q": q,

        "part_types_selected": part_types,
        "region": region,
        "city": city,
        "condition": condition,
        "color": color,
        "price_min": price_min_raw,
        "price_max": price_max_raw,
        "has_images": has_images,
        "sort": sort,
        "has_filters": has_filters,

        "facet_part_types": facet_part_types,
        "facet_conditions": facet_conditions,
        "facet_colors": facet_colors,
        "facet_cities": facet_cities,
        "facet_regions": facet_regions,
        "price_bounds": price_bounds,
    })



@login_required
def warehouse_view(request):
    q_raw = (request.GET.get("q") or "").strip()

    # совместимость со старыми полями
    brand_in = (request.GET.get("brand") or "").strip()
    model_in = (request.GET.get("model") or "").strip()

    # ✅ новый фильтр (мульти-чекбоксы)
    part_types = [p.strip() for p in request.GET.getlist("part_type") if p.strip()]

    # если где-то остались старые ссылки ?part_type=... (одним значением) — поддержим
    legacy_part_type = (request.GET.get("part_type") or "").strip()
    if legacy_part_type and legacy_part_type not in part_types:
        part_types.append(legacy_part_type)

    brand, model, single_term, term, q = split_search_q(q_raw, brand_in, model_in)

    qs = (
        Part.objects
        .filter(user=request.user)
        .prefetch_related("images")
    )
    qs = _annotate_search_fields(qs)

    # метрики
    total_views = (
        Part.objects
        .filter(user=request.user)
        .aggregate(total_views=Sum("views"))
        .get("total_views") or 0
    )

    brand_l = (brand or "").lower()
    model_l = (model or "").lower()
    term_l = (term or "").lower()

    model_norm = normalize_compact(model)
    term_norm = normalize_compact(term)

    note_pat_term = note_regex_from_text(term) if term else None
    note_pat_model = note_regex_from_text(model) if model else None

    # --- тот же поиск, что и в search() ---
    if single_term and term:
        base_q = (
            Q(normalized_model__icontains=term_norm) |
            Q(lower_model__icontains=term_l) |
            Q(lower_chip_label__icontains=term_l) |
            Q(normalized_part_number=term_norm) |
            Q(lower_brand__icontains=term_l)
        )
        if note_pat_term:
            base_q = base_q | Q(note__iregex=note_pat_term)

        qs = qs.filter(base_q)

        whens = [
            When(normalized_part_number=term_norm, then=Value(120)),
            When(normalized_model=term_norm, then=Value(95)),
            When(lower_model=term_l, then=Value(90)),
            When(normalized_model__icontains=term_norm, then=Value(75)),
            When(lower_model__icontains=term_l, then=Value(70)),
            When(lower_chip_label__icontains=term_l, then=Value(60)),
            When(lower_brand__icontains=term_l, then=Value(40)),
        ]
        if note_pat_term:
            whens.append(When(note__iregex=note_pat_term, then=Value(25)))

        qs = qs.annotate(match_priority=Case(*whens, default=Value(0), output_field=IntegerField()))

    else:
        if brand:
            qs = qs.filter(lower_brand__icontains=brand_l)

        if model:
            model_q = (
                Q(normalized_model__icontains=model_norm) |
                Q(lower_model__icontains=model_l) |
                Q(lower_chip_label__icontains=model_l) |
                Q(normalized_part_number=model_norm)
            )
            if note_pat_model:
                model_q = model_q | Q(note__iregex=note_pat_model)

            qs = qs.filter(model_q)

        whens = []
        if model:
            whens.extend([
                When(normalized_part_number=model_norm, then=Value(115)),
                When(normalized_model=model_norm, then=Value(85)),
                When(lower_model=model_l, then=Value(80)),
                When(normalized_model__icontains=model_norm, then=Value(65)),
                When(lower_model__icontains=model_l, then=Value(60)),
                When(lower_chip_label__icontains=model_l, then=Value(55)),
            ])
            if brand:
                whens.insert(1, When(lower_brand=brand_l, normalized_model=model_norm, then=Value(100)))
                whens.insert(2, When(lower_brand=brand_l, lower_model=model_l, then=Value(95)))
            if note_pat_model:
                whens.append(When(note__iregex=note_pat_model, then=Value(25)))

        if brand:
            whens.append(When(lower_brand=brand_l, then=Value(30)))

        qs = qs.annotate(match_priority=Case(*whens, default=Value(0), output_field=IntegerField()))

    # ✅ фасеты по типам (после поиска, но ДО применения выбранных типов)
    facets_base = qs
    facet_part_types = list(
        facets_base.exclude(part_type__isnull=True).exclude(part_type="")
        .values("part_type").annotate(count=Count("id")).order_by("part_type")
    )

    # ✅ применяем выбранные типы (точное совпадение по значению чекбокса)
    if part_types:
        qs = qs.filter(part_type__in=part_types)

    qs = qs.order_by("-match_priority", "-created_at")

    total_parts = qs.count()
    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get("page"))

    # ✅ base_qs для пагинации (сохраняет q + part_type + всё остальное)
    get_copy = request.GET.copy()
    get_copy.pop("page", None)
    base_qs = get_copy.urlencode()

    devices = (
        Part.objects.filter(user=request.user)
        .values_list("device", flat=True).distinct().order_by("device")
    )
    brands = (
        Part.objects.filter(user=request.user)
        .values_list("brand", flat=True).distinct().order_by("brand")
    )

    return render(request, "warehouse/warehouse.html", {
        "page_obj": page_obj,
        "base_qs": base_qs,

        "query": q,  # строка поиска
        "total_parts": total_parts,
        "total_views": total_views,

        # фильтр part_type
        "facet_part_types": facet_part_types,
        "part_types_selected": part_types,

        # твои кнопки
        "devices": devices,
        "brands": brands,
    })
def logout_view(request):
    logout(request)
    return redirect('home')  # Перенаправление на главную страницу после выхода





@login_required
def edit_part(request, part_id):
    part = get_object_or_404(Part, id=part_id)

    if request.method == 'POST':

        part.device = request.POST.get('device')
        part.brand = request.POST.get('brand')
        part.model = request.POST.get('model')
        part.part_type = request.POST.get('part_type')
        part.color = request.POST.get('color')
        part.quantity = request.POST.get('quantity')
        part.price = request.POST.get('price')
        part.note = request.POST.get('note')
        part.save()

        # Обрабатываем новые изображения (если загружены)
        if 'images' in request.FILES:
            images = request.FILES.getlist('images')
            for image in images:
                if part.images.count() < 5:
                    part.images.create(image=image)

        return redirect('warehouse')

    return render(request, 'warehouse/edit_part.html', {'part': part})



@login_required
def delete_part(request, part_id):
    part = get_object_or_404(Part, id=part_id, user=request.user)

    if request.method == 'POST':
        # 1) Пытаемся удалить сообщение в Telegram (если есть сохранённый message_id)
        try:
            delete_telegram_message_for_part(part.id)
        except Exception as e:
            # Не ломаем UX, просто лог
            print(f"[telegram] delete on part remove failed: {e}")

        # 2) Удаляем саму запчасть
        part.delete()
        return redirect('warehouse')

    return render(request, 'warehouse/delete_part.html', {'part': part})



@login_required(login_url='login')
def profile(request):
    return render(request, 'warehouse/profile.html')


@login_required(login_url='login')
def warehouse(request):
    return render(request, 'warehouse/warehouse.html')




@login_required
def export_excel(request):
    # Создаем новый Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Запчасти"

    # Заголовки столбцов
    headers = ["Устройство", "Бренд", "Модель", "Тип запчасти", "Цвет", "Количество (шт.)", "Цена (руб.)"]
    ws.append(headers)

    # Применяем жирный шрифт к заголовкам
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font

    # Получаем данные из модели, сортируем по устройству, бренду и модели
    user_parts = Part.objects.filter(user=request.user).order_by('device', 'brand', 'model')

    # Определяем цвета заливки
    phone_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Светло-голубой
    tablet_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Светло-зелёный
    watch_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")   # Светло-жёлтый
    separator_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Светло-серый

    row_num = 2  # Начинаем со второй строки (первая — заголовки)

    # Группируем данные только по устройству
    for device, device_parts in groupby(user_parts, lambda x: x.device):
        # Определяем цвет заливки для столбца "Устройство"
        if device == "Телефон":
            fill = phone_fill
        elif device == "Планшет":
            fill = tablet_fill
        elif device == "Смарт-часы":
            fill = watch_fill
        else:
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Белый по умолчанию

        # Записываем все запчасти для текущего устройства
        for part in device_parts:
            ws.append([
                part.device,
                part.brand,
                part.model,
                part.part_type,
                part.color or "Не указан",
                part.quantity,
                part.price
            ])

            # Применяем заливку только к ячейке в столбце "Устройство" (первый столбец)
            device_cell = ws.cell(row=row_num, column=1)
            device_cell.fill = fill

            # Форматируем столбцы "Количество" и "Цена"
            quantity_cell = ws.cell(row=row_num, column=6)
            quantity_cell.number_format = '#,##0 "шт."'
            price_cell = ws.cell(row=row_num, column=7)
            price_cell.number_format = '#,##0 "руб."'

            row_num += 1

        # Добавляем разделительную строку после группы устройств
        for col_num in range(1, 8):
            separator_cell = ws.cell(row=row_num, column=col_num)
            separator_cell.fill = separator_fill
        row_num += 1

    # Настройка ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Создаем HTTP ответ с Excel файлом
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=parts.xlsx'
    wb.save(response)

    return response

@login_required
def get_brands_for_device(request):
    device = request.GET.get('device')
    brands = Part.objects.filter(user=request.user, device=device).values_list('brand', flat=True).distinct()
    return JsonResponse({'brands': list(brands)})


@login_required
def filter_parts(request):
    device = request.GET.get('device')
    brand = request.GET.get('brand')
    model = request.GET.get('model')
    part_type = request.GET.get('part_type')

    # Фильтрация запчастей по устройству, бренду, модели и типу
    parts = Part.objects.filter(user=request.user, device=device, brand=brand)
    if model:
        parts = parts.filter(model__icontains=model)
    if part_type:
        parts = parts.filter(part_type__icontains=part_type)

    # Подготовка данных для ответа
    parts_data = [
        {
            'id': part.id,
            'device': part.device,
            'brand': part.brand,
            'model': part.model,
            'part_type': part.part_type,
            'color': part.color,
            'quantity': part.quantity,
            'price': part.price,
            'image': part.image.url if part.image else None
        }
        for part in parts
    ]

    return JsonResponse({'parts': parts_data})


def part_detail(request, part_id):
    part = get_object_or_404(Part, id=part_id)

    if request.user.is_authenticated and request.user != part.user:
        part.views += 1
        part.last_viewed = now()
        part.save(update_fields=['views', 'last_viewed'])

    is_bookmarked = False
    if request.user.is_authenticated:
        is_bookmarked = request.user.bookmarks.filter(part=part).exists()

    return render(request, 'warehouse/part_detail.html', {
        'part': part,
        'is_bookmarked': is_bookmarked,
    })


def add_part_success(request):
    return render(request, 'warehouse/add_part_success.html')


def delete_image(request, image_id):
    if request.method == 'POST' and request.user.is_authenticated:
        # Получаем изображение по ID
        image = get_object_or_404(PartImage, id=image_id)

        # Проверяем, принадлежит ли изображение текущему пользователю (если нужно)
        # if image.part.user != request.user:
        #     return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

        # Удаляем изображение
        image.delete()

        # Возвращаем ответ
        return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)


@login_required
def add_image(request):
    if request.method == 'POST':
        part_id = request.POST.get('part_id')
        images = request.FILES.getlist('image')

        part = get_object_or_404(Part, id=part_id)

        if part.images.count() + len(images) > 5:
            return JsonResponse({'status': 'error', 'message': 'Максимальное количество изображений — 5.'})

        image_responses = []
        for image in images:
            compressed = compress_image(image)
            watermarked = add_watermark_to_image(compressed)
            part_image = PartImage(part=part, image=watermarked)
            part_image.save()
            image_responses.append({
                'id': part_image.id,
                'url': part_image.image.url,
            })

        return JsonResponse({'status': 'success', 'images': image_responses})
    return JsonResponse({'status': 'error', 'message': 'Неверный метод запроса.'})


@login_required
def get_devices(request):
    devices = Part.objects.filter(user=request.user).values_list('device', flat=True).distinct().order_by('device')
    return JsonResponse({'devices': list(devices)})

@login_required
def get_brands(request):
    device = request.GET.get('device')
    brands = Part.objects.filter(user=request.user, device=device).values_list('brand', flat=True).distinct().order_by('brand')
    return JsonResponse({'brands': list(brands)})

@login_required
def get_models(request):
    device = request.GET.get('device')
    brand = request.GET.get('brand')
    models = Part.objects.filter(user=request.user, device=device, brand=brand).values_list('model', flat=True).distinct().order_by('model')
    return JsonResponse({'models': list(models)})

@login_required
def get_part_types(request):
    device = request.GET.get('device')
    brand = request.GET.get('brand')
    model = request.GET.get('model')
    part_types = Part.objects.filter(user=request.user, device=device, brand=brand, model=model).values_list('part_type', flat=True).distinct().order_by('part_type')
    return JsonResponse({'part_types': list(part_types)})

@login_required
def get_parts(request):
    device = request.GET.get('device')
    brand = request.GET.get('brand')
    model = request.GET.get('model')
    part_type = request.GET.get('part_type')

    filters = {'user': request.user}
    if device:
        filters['device'] = device
    if brand:
        filters['brand'] = brand
    if model:
        filters['model'] = model
    if part_type:
        filters['part_type'] = part_type

    parts = Part.objects.filter(**filters).prefetch_related('images')
    parts_data = [{
        'id': part.id,
        'device': part.device,
        'brand': part.brand,
        'model': part.model,
        'part_type': part.part_type,
        'part_number': part.part_number or '',  # Добавляем part_number
        'color': part.color,
        'quantity': part.quantity,
        'price': str(part.price),  # Преобразуем Decimal в строку для JSON
        'note': part.note,
        'condition': part.condition,
        'images': [{'image_url': image.image.url} for image in part.images.all()]
    } for part in parts]

    return JsonResponse({'parts': parts_data})


def base_view(request):
    # Подсчёт количества пользователей и запчастей
    user_count = User.objects.count()
    part_count = Part.objects.count()

    return render(request, 'warehouse/base.html', {
        'user_count': user_count,
        'part_count': part_count,
    })


# Путь к JSON файлу
# DATA_FILE = Path(__file__).resolve().parent.parent / "json_manager" / "г.json"

DATA_FILE = Path(__file__).resolve().parent / "data.json"

def load_data():
    with open(DATA_FILE, 'r', encoding='utf-8') as file:
        return json.load(file)


def get_dynamic_data(request):
    """Универсальное представление для динамического обновления данных."""
    data = load_data()  # Загружаем данные из JSON
    device = request.GET.get("device")
    brand = request.GET.get("brand")
    part_type = request.GET.get("part_type")
    response = {}

    if device:
        response["brands"] = data.get(device, {}).get("brands", [])
        response["part_types"] = data.get(device, {}).get("part_types", [])

    if device and brand:
        response["models"] = data.get(device, {}).get("models", {}).get(brand, [])

    if device and part_type:
        response["colors"] = data.get(device, {}).get("colors", {}).get(part_type, [])
        response["conditions"] = data.get(device, {}).get("conditions", [])

    return JsonResponse(response)



# Регистрация pillow-heif для поддержки HEIC-формата
try:
    from pillow_heif import register_heif_opener
    register_heif_opener()
except ImportError:
    # Если библиотека не установлена, HEIC-формат не будет поддерживаться
    pass


import os
import uuid
import time

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from .forms import PartForm, PartImageFormSet
from .models import Part, PartImage
from tariff.utils import check_parts_limit, check_image_limit
from .utils import compress_image
from warehouse.utils import add_watermark_to_image
from .telegram_utils import send_new_part_notification

# куда складываем временные картинки
ADD_PART_TMP_DIR = "tmp/add_part"


def _cleanup_draft_files(file_names: list[str]):
    for name in file_names or []:
        try:
            default_storage.delete(name)
        except Exception:
            pass


def _save_uploaded_files_to_tmp(request, draft_token: str) -> list[str]:
    """
    Сохраняем ВСЕ файлы из request.FILES во временное хранилище.
    В сессию кладём только имена файлов (строки).
    """
    saved = []
    user_id = request.user.id

    # request.FILES может содержать несколько ключей (form-0-image, form-1-image, ...)
    for field_name in request.FILES:
        for f in request.FILES.getlist(field_name):
            # уникальное имя
            safe_name = f"{uuid.uuid4().hex}_{os.path.basename(getattr(f, 'name', 'upload'))}"
            path = f"{ADD_PART_TMP_DIR}/{user_id}/{draft_token}/{safe_name}"
            saved_name = default_storage.save(path, f)
            saved.append(saved_name)

    return saved


def _get_draft_from_session(request):
    draft = request.session.get("add_part_draft")
    if not isinstance(draft, dict):
        return None
    return draft


def _pop_draft_from_session(request):
    draft = request.session.pop("add_part_draft", None)
    if not isinstance(draft, dict):
        return None
    return draft


@login_required
def add_part(request):
    """
    Добавление запчасти + обработка дубля:
    - если дубль найден -> показываем confirm_add_part.html
      и сохраняем фото во временную папку (в session только пути)
    - если confirm_add -> создаём новую запчасть и переносим фото из tmp
    """

    profile = getattr(request.user, "profile", None)
    if not profile or not profile.city or not profile.phone:
        messages.error(
            request,
            "Перед началом создания склада, укажите, пожалуйста, город и номер телефона в вашем профиле."
        )
        return redirect("profile")

    # --- GET ---
    if request.method == "GET":
        # (необязательно) чистим очень старый черновик, если завис
        draft = _get_draft_from_session(request)
        if draft and isinstance(draft.get("created_ts"), (int, float)):
            if time.time() - draft["created_ts"] > 60 * 60:  # 1 час
                _cleanup_draft_files(draft.get("tmp_files", []))
                request.session.pop("add_part_draft", None)

        form = PartForm()
        formset = PartImageFormSet(queryset=PartImage.objects.none())
        return render(request, "warehouse/add_part.html", {"form": form, "formset": formset})

    # --- POST ---

    # 0) пользователь нажал "Отмена" на confirm-странице
    if request.POST.get("discard_draft") == "1":
        draft = _pop_draft_from_session(request)
        if draft:
            _cleanup_draft_files(draft.get("tmp_files", []))
        return redirect("add_part")

    # 1) пользователь нажал "Добавить как новую" на confirm-странице
    if request.POST.get("confirm_add") == "1":
        draft = _pop_draft_from_session(request)
        if not draft:
            messages.error(request, "Черновик добавления устарел. Пожалуйста, попробуйте ещё раз.")
            return redirect("add_part")

        form_data = draft.get("form_data") or {}
        tmp_files = draft.get("tmp_files") or []

        send_to_telegram = (form_data.get("send_to_telegram") == "on")

        # валидируем заново (без FILES, фото приделаем вручную ниже)
        form = PartForm(form_data)

        if not form.is_valid():
            _cleanup_draft_files(tmp_files)
            messages.error(request, "Данные формы некорректны. Заполните форму снова.")
            formset = PartImageFormSet(queryset=PartImage.objects.none())
            return render(request, "warehouse/add_part.html", {"form": form, "formset": formset})

        # лимит картинок на free проверим ещё раз
        if request.user.profile.tariff == "free" and tmp_files:
            if check_image_limit(request.user, len(tmp_files)):
                _cleanup_draft_files(tmp_files)
                messages.error(request, "На бесплатном тарифе можно загрузить только 1 изображение для запчасти.")
                formset = PartImageFormSet(queryset=PartImage.objects.none())
                return render(request, "warehouse/add_part.html", {"form": form, "formset": formset})

        # микросхема + маркировка
        part_type_value = form.cleaned_data["part_type"]
        chip_marking = (form_data.get("chip_marking") or "").strip()
        if part_type_value == "Микросхема" and chip_marking:
            form.cleaned_data["part_type"] = f"{part_type_value} {chip_marking}"

        with transaction.atomic():
            new_part = form.save(commit=False)
            new_part.user = request.user
            new_part.save()

            # переносим tmp фото -> PartImage
            # ограничим 5 фото на всякий случай
            for name in tmp_files[:5]:
                try:
                    with default_storage.open(name, "rb") as f:
                        content = f.read()
                    original_name = os.path.basename(name).split("_", 1)[-1] or "image.jpg"
                    cf = ContentFile(content, name=original_name)

                    compressed = compress_image(cf)
                    watermarked = add_watermark_to_image(compressed)
                    PartImage.objects.create(part=new_part, image=watermarked)
                finally:
                    # удаляем временный файл в любом случае
                    try:
                        default_storage.delete(name)
                    except Exception:
                        pass

            if send_to_telegram:
                transaction.on_commit(lambda: send_new_part_notification(new_part, request=request))

        return render(request, "warehouse/success.html", {"message": "Запчасть успешно добавлена!"})

    # 2) обычный первый POST с формы
    if check_parts_limit(request.user):
        tariff = request.user.profile.tariff
        if tariff == "free":
            msg = "Вы достигли лимита в 30 запчастей для бесплатного тарифа. Для добавления новых запчастей обновите тариф."
        elif tariff == "lite":
            msg = "Вы достигли лимита в 500 запчастей для тарифа Lite. Для добавления новых запчастей обновите тариф."
        elif tariff == "standard":
            msg = "Вы достигли лимита в 2000 запчастей для тарифа Стандарт. Для добавления новых запчастей обновите тариф."
        elif tariff == "standard2":
            msg = "Вы достигли лимита в 7000 запчастей для тарифа Стандарт 2. Для добавления новых запчастей обновите тариф."
        elif tariff == "standard3":
            msg = "Вы достигли лимита в 15000 запчастей для тарифа Стандарт 3. Для добавления новых запчастей обновите тариф."
        else:
            msg = "Лимит запчастей достигнут."
        messages.error(request, msg)
        return redirect("profile")

    send_to_telegram = request.POST.get("send_to_telegram") == "on"

    form = PartForm(request.POST, request.FILES)
    formset = PartImageFormSet(request.POST, request.FILES, queryset=PartImage.objects.none())

    if not (form.is_valid() and formset.is_valid()):
        messages.error(request, "Ошибка валидации формы. Проверьте введенные данные.")
        return render(request, "warehouse/add_part.html", {"form": form, "formset": formset})

    # лимит картинок free
    images_forms = [f for f in formset if f.cleaned_data and f.cleaned_data.get("image")]
    if request.user.profile.tariff == "free" and check_image_limit(request.user, len(images_forms)):
        messages.error(request, "На бесплатном тарифе можно загрузить только 1 изображение для запчасти.")
        return render(request, "warehouse/add_part.html", {"form": form, "formset": formset})

    # микросхема + маркировка
    part_type_value = form.cleaned_data["part_type"]
    chip_marking = (request.POST.get("chip_marking") or "").strip()
    if part_type_value == "Микросхема" and chip_marking:
        form.cleaned_data["part_type"] = f"{part_type_value} {chip_marking}"

    device = form.cleaned_data["device"]
    brand = form.cleaned_data["brand"]
    model = form.cleaned_data["model"]
    part_type = form.cleaned_data["part_type"]
    condition = form.cleaned_data["condition"]
    color = form.cleaned_data.get("color") or None

    existing_part = find_existing_part_for_user(
        user=request.user,
        device=device,
        brand=brand,
        model=model,
        part_type=part_type,
        condition=condition,
        color=color,
    )

    # если дубль — показываем confirm и сохраняем фото во временное хранилище
    if existing_part:
        draft_token = uuid.uuid4().hex
        tmp_files = _save_uploaded_files_to_tmp(request, draft_token)

        # ВАЖНО: в session только JSON-данные (строки/числа/списки)
        request.session["add_part_draft"] = {
            "created_ts": int(time.time()),
            "form_data": request.POST.dict(),   # достаточно для PartForm
            "tmp_files": tmp_files,
            "existing_part_id": existing_part.id,
        }

        return render(request, "warehouse/confirm_add_part.html", {
            "existing_part": existing_part,
            "tmp_images_count": len(tmp_files),
        })

    # если не дубль — обычное сохранение
    with transaction.atomic():
        new_part = form.save(commit=False)
        new_part.user = request.user
        new_part.save()

        for img_form in formset:
            if img_form.cleaned_data and img_form.cleaned_data.get("image"):
                image_obj = img_form.save(commit=False)
                compressed = compress_image(image_obj.image)
                watermarked = add_watermark_to_image(compressed)
                image_obj.image = watermarked
                image_obj.part = new_part
                image_obj.save()

        if send_to_telegram:
            transaction.on_commit(lambda: send_new_part_notification(new_part, request=request))

    return render(request, "warehouse/success.html", {"message": "Запчасть успешно добавлена!"})

def get_regions_and_cities(request):
    file_path = os.path.join(settings.BASE_DIR, 'static/json/belarus_regions_and_cities.json')
    with open(file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)
    return JsonResponse(data)


def user_parts_list(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user_parts = Part.objects.filter(user_id=user_id, quantity__gt=0).order_by('device', 'brand', 'model')

    grouped_parts = {}
    for device, device_parts in groupby(user_parts, lambda x: x.device):
        grouped_parts[device] = {}
        for brand, brand_parts in groupby(device_parts, lambda x: x.brand):
            grouped_parts[device][brand] = list(brand_parts)

    return render(request, 'warehouse/user_parts.html', {
        'grouped_parts': grouped_parts,
        'viewed_user': user,
    })

@login_required
def user_parts(request, user_id, template_name='warehouse/user_parts.html'):
    user = get_object_or_404(User, id=user_id)
    user_parts = Part.objects.filter(user_id=user_id).order_by('device', 'brand', 'model')

    grouped_parts = {}
    for device, device_parts in groupby(user_parts, lambda x: x.device):
        grouped_parts[device] = {}
        for brand, brand_parts in groupby(device_parts, lambda x: x.brand):
            grouped_parts[device][brand] = list(brand_parts)

    return render(request, template_name, {
        'grouped_parts': grouped_parts,
        'viewed_user': user,
        'viewed_user_full_name': user.profile.full_name,
    })


@login_required
def delete_all_parts_page(request):
    if request.method == 'POST':
        user_parts = Part.objects.filter(user=request.user)

        for part in user_parts:
            for image in part.images.all():
                if image.image:  # Проверяем, есть ли файл
                    image_path = os.path.join(settings.MEDIA_ROOT, str(image.image))
                    if os.path.exists(image_path):
                        os.remove(image_path)  # Удаляем файл изображения

                image.delete()  # Удаляем запись из базы

        user_parts.delete()  # Удаляем запчасти

        messages.success(request, 'Все запчасти и их изображения были успешно удалены.')
        return redirect('warehouse')

    return render(request, 'warehouse/delete_all_parts.html')


def about_project(request):
    return render(request, 'warehouse/about_project.html')


def search_user_parts(request, user_id):
    viewed_user = get_object_or_404(User, id=user_id)
    parts = Part.objects.filter(user=viewed_user, quantity__gt=0)

    # Обработка параметров поиска
    query = request.GET.get('query', '').strip()
    if query:
        parts = parts.filter(
            Q(model__icontains=query) |  # Используем поле model
            Q(part_type__icontains=query)
        )

    # Группировка запчастей
    grouped_parts = {}
    for part in parts:
        grouped_parts.setdefault(part.device, {}).setdefault(part.brand, []).append(part)

    return render(request, 'warehouse/user_parts.html', {
        'viewed_user': viewed_user,
        'grouped_parts': grouped_parts,
    })


def user_agreement_view(request):
    return render(request, 'warehouse/user_agreement.html')



def privacy_policy_view(request):
    return render(request, 'warehouse/privacy_policy.html')


@login_required
def analytics_view(request):
    user = request.user
    today = timezone.now().date()
    month_ago = today - timedelta(days=30)

    # Общие метрики
    total_views = Part.objects.filter(user=user).aggregate(Sum('views'))['views__sum'] or 0
    daily_views = Part.objects.filter(
        user=user,
        last_viewed__date=today
    ).aggregate(Sum('views'))['views__sum'] or 0
    monthly_views = Part.objects.filter(
        user=user,
        last_viewed__date__gte=month_ago
    ).aggregate(Sum('views'))['views__sum'] or 0

    # Данные для графика (просмотры по дням за последние 30 дней)
    views_by_day = (
        Part.objects.filter(user=user, last_viewed__date__gte=month_ago)
        .annotate(date=TruncDate('last_viewed', output_field=DateField()))
        .values('date')
        .annotate(views=Sum('views'))
        .order_by('date')
    )

    # Форматируем данные для Chart.js
    labels = []
    data = []
    current_date = month_ago
    while current_date <= today:
        labels.append(current_date.strftime('%d %b'))
        # Ищем просмотры для текущей даты
        views = next((item['views'] for item in views_by_day if item['date'] == current_date), 0)
        data.append(views or 0)
        current_date += timedelta(days=1)

    context = {
        'total_views': total_views,
        'daily_views': daily_views,
        'monthly_views': monthly_views,
        'today': today,
        'month_ago': month_ago,
        'chart_labels': labels,  # Даты для осей X
        'chart_data': data,     # Просмотры для осей Y
    }
    return render(request, 'warehouse/analytics.html', context)



@login_required
def add_full_device_parts(request):
    """
    Массовое добавление запчастей для одного устройства.

    ВАЖНО:
    - Проверка на дубликаты и слияние количеств происходит на фронтенде,
      в модальном окне, через AJAX:
        * ajax_check_full_device_part_duplicate
        * ajax_merge_full_device_part
    - Эта вьюха на POST просто создаёт НОВЫЕ запчасти по всем частям,
      у которых стоит чекбокс parts-*-selected == "on" И parts-*-merged == "0".
      Для тех, которые были слиты с уже существующими, JS ставит merged = "1",
      и мы их пропускаем.
    """

    profile = getattr(request.user, "profile", None)
    if not profile or not profile.city or not profile.phone:
        messages.error(
            request,
            "Перед началом создания склада, укажите, пожалуйста, город и номер телефона в вашем профиле.",
        )
        return redirect("profile")

    # -------- GET: просто отрисовать пустую форму --------
    if request.method == "GET":
        return render(
            request,
            "warehouse/add_full_device_parts.html",
            {
                "device": "",
                "brand": "",
                "model": "",
            },
        )

    # -------- POST: массовое сохранение выбранных запчастей --------

    # 1. Проверка лимита по количеству запчастей
    if check_parts_limit(request.user):
        tariff = request.user.profile.tariff
        if tariff == "free":
            error_message = "Вы достигли лимита в 30 запчастей для бесплатного тарифа. Для добавления новых запчастей обновите тариф."
        elif tariff == "lite":
            error_message = "Вы достигли лимита в 500 запчастей для тарифа Lite. Для добавления новых запчастей обновите тариф."
        elif tariff == "standard":
            error_message = "Вы достигли лимита в 2000 запчастей для тарифа Стандарт. Для добавления новых запчастей обновите тариф."
        elif tariff == "standard2":
            error_message = "Вы достигли лимита в 7000 запчастей для тарифа Стандарт 2. Для добавления новых запчастей обновите тариф."
        elif tariff == "standard3":
            error_message = "Вы достигли лимита в 15000 запчастей для тарифа Стандарт 3. Для добавления новых запчастей обновите тариф."
        else:
            error_message = "Лимит запчастей достигнут."
        messages.error(request, error_message)
        return redirect("profile")

    # 2. Базовые данные устройства (БЕЗ общего цвета и БЕЗ общего send_to_telegram)
    device = (request.POST.get("device") or "").strip()
    brand = (request.POST.get("brand") or "").strip()
    model = (request.POST.get("model") or "").strip()

    if not (device and brand and model):
        messages.error(request, "Заполните устройство, бренд и модель.")
        return render(
            request,
            "warehouse/add_full_device_parts.html",
            {
                "device": device,
                "brand": brand,
                "model": model,
            },
        )

    # 3. Общее количество "пинов" (типов запчастей)
    try:
        total_parts = int(request.POST.get("parts_total", "0"))
    except ValueError:
        total_parts = 0

    if total_parts <= 0:
        messages.error(request, "Не выбрано ни одной запчасти.")
        return render(
            request,
            "warehouse/add_full_device_parts.html",
            {
                "device": device,
                "brand": brand,
                "model": model,
            },
        )

    created_parts = []

    # 4. Сохраняем всё одной транзакцией
    with transaction.atomic():
        for i in range(total_parts):
            # чекбокс "Есть"
            if request.POST.get(f"parts-{i}-selected") != "on":
                continue

            # если эта запчасть уже была объединена (merged) с существующей — пропускаем
            merged_flag = request.POST.get(f"parts-{i}-merged", "0")
            if merged_flag == "1":
                continue

            part_type = (request.POST.get(f"parts-{i}-part_type") or "").strip()
            if not part_type:
                continue

            chip_marking = (request.POST.get(f"parts-{i}-chip_marking") or "").strip()

            # для микросхем добавляем маркировку в тип (как в одиночном add_part)
            if part_type == "Микросхема" and chip_marking:
                part_type_for_save = f"{part_type} {chip_marking}"
            else:
                part_type_for_save = part_type

            quantity_raw = request.POST.get(f"parts-{i}-quantity") or "1"
            price_raw    = request.POST.get(f"parts-{i}-price")    or "0"
            color        = (request.POST.get(f"parts-{i}-color") or "").strip() or None
            condition    = (request.POST.get(f"parts-{i}-condition") or "").strip()
            note         = (request.POST.get(f"parts-{i}-note") or "").strip()

            # флаг "отправлять эту запчасть в Telegram"
            send_to_telegram_flag = request.POST.get(f"parts-{i}-send_tg") == "1"

            if not condition:
                condition = "оригинал б/у"

            try:
                quantity = int(quantity_raw)
            except ValueError:
                quantity = 1

            price = price_raw  # DecimalField сам приведёт строку

            # файл картинки: максимум 1 на запчасть
            image_file = request.FILES.get(f"parts-{i}-image")

            # проверка лимита картинок для бесплатного тарифа
            if image_file and request.user.profile.tariff == "free":
                if check_image_limit(request.user, 1):
                    messages.warning(
                        request,
                        f"Для запчасти «{part_type_for_save}» превышен лимит изображений на бесплатном тарифе. "
                        f"Она будет создана без фото.",
                    )
                    image_file = None

            # создаём Part
            new_part = Part(
                user=request.user,
                device=device,
                brand=brand,
                model=model,
                part_type=part_type_for_save,
                quantity=quantity,
                price=price,
                color=color,
                condition=condition,
                note=note,
            )
            new_part.save()

            # если есть картинка — сжимаем + водяной знак
            if image_file:
                compressed  = compress_image(image_file)
                watermarked = add_watermark_to_image(compressed)
                PartImage.objects.create(part=new_part, image=watermarked)

            created_parts.append(new_part)

            # Отправка в Telegram именно для ЭТОЙ запчасти, если включен флаг
            if send_to_telegram_flag:
                transaction.on_commit(
                    lambda part=new_part: send_new_part_notification(part, request=request)
                )

    return render(
        request,
        "warehouse/success.html",
        {
            "message": f"Запчасти ({len(created_parts)} шт.) успешно добавлены!",
        },
    )



@login_required
def ajax_check_full_device_part_duplicate(request):
    if request.method != "POST":
        return JsonResponse({"error": "method_not_allowed"}, status=405)

    device    = request.POST.get("device")
    brand     = request.POST.get("brand")
    model     = request.POST.get("model")
    part_type = request.POST.get("part_type")
    condition = request.POST.get("condition")
    color     = request.POST.get("color")

    if not (device and brand and model and part_type):
        return JsonResponse({"error": "missing_fields"}, status=400)

    existing = find_existing_part_for_user(
        user=request.user,
        device=device,
        brand=brand,
        model=model,
        part_type=part_type,
        condition=condition,
        color=color,
    )

    if not existing:
        return JsonResponse({"exists": False})

    return JsonResponse({
        "exists": True,
        "part_id": existing.id,
        "quantity": existing.quantity,
        "price": str(existing.price) if existing.price is not None else "",
        "edit_url": reverse("edit_part", args=[existing.id]),
    })


@login_required
def ajax_merge_full_device_part(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'method_not_allowed'}, status=405)

    try:
        part_id = int(request.POST.get('part_id', '0'))
    except ValueError:
        return JsonResponse({'error': 'bad_part_id'}, status=400)

    try:
        add_quantity = int(request.POST.get('add_quantity', '0'))
    except ValueError:
        add_quantity = 0

    new_price = (request.POST.get('price') or '').strip()

    part = get_object_or_404(Part, id=part_id, user=request.user)

    if add_quantity > 0:
        part.quantity = (part.quantity or 0) + add_quantity

    if new_price:
        try:
            part.price = new_price
        except Exception:
            pass

    part.save(update_fields=['quantity', 'price'])

    return JsonResponse({
        'ok': True,
        'new_quantity': part.quantity,
        'price': str(part.price) if part.price is not None else '',
    })



def find_existing_part_for_user(user, device, brand, model, part_type, condition, color):
    """
    Унифицированный поиск "такой же" запчасти.

    - Сравнение по user/device/brand/model/part_type/condition — через __iexact.
    - По цвету:
        * если color не передан (None или пустая строка) — цвет игнорируем
          и ищем среди всех цветов;
        * если color есть — фильтруем по нему (__iexact).
    """

    device    = (device or "").strip()
    brand     = (brand or "").strip()
    model     = (model or "").strip()
    part_type = (part_type or "").strip()
    condition = (condition or "").strip() or "оригинал б/у"
    color     = (color or "").strip() or None

    qs = Part.objects.filter(
        user=user,
        device__iexact=device,
        brand__iexact=brand,
        model__iexact=model,
        part_type__iexact=part_type,
        condition__iexact=condition,
    )

    if color is not None:
        qs = qs.filter(color__iexact=color)

    return qs.first()
